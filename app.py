import streamlit as st
import pandas as pd
import os
import re
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import xlrd


#  https://github.com/luiginoIori/4I_Capital.git

def is_date_formatted(cell):
    """Verifica se uma célula está formatada como data"""
    if cell.value is None:
        return False
    
    # Verifica se é datetime
    if isinstance(cell.value, datetime):
        return True
    
    # Verifica se é string que parece data no formato ##/##/####
    if isinstance(cell.value, str):
        date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
        return bool(re.match(date_pattern, cell.value))
    
    # Verifica se é número de data do Excel
    if isinstance(cell.value, (int, float)) and cell.number_format:
        # Formatos de data comuns no Excel
        date_formats = ['dd/mm/yyyy', 'd/m/yyyy', 'mm/dd/yyyy', 'm/d/yyyy', 'yyyy-mm-dd']
        return any(fmt in cell.number_format.lower() for fmt in date_formats)
    
    return False

def process_sicred_files(arquivos):
    """Processa todos os arquivos Sicred de 2025"""
    arq_data =[]
    # Lista arquivos Sicred de 2025
    arquivos_sicred_2025 = []
    for arquivo in arquivos:
        if "2025" in arquivo and "Sicred" in arquivo and (arquivo.endswith('.xls') or arquivo.endswith('.xlsx')):
            arquivos_sicred_2025.append(arquivo)
    
      
    for arquivo in arquivos_sicred_2025:
        caminho_arquivo = arquivo
        
        try:
            # Para arquivos .xls antigos, usa xlrd
            if arquivo.endswith('.xls'):
                workbook_xlrd = xlrd.open_workbook(caminho_arquivo)
                sheet_xlrd = workbook_xlrd.sheet_by_index(0)
                
                
                sair = False
                # Processa linha por linha
                for row_num in range(sheet_xlrd.nrows):
                    try:
                        # Verifica se a célula na coluna A (índice 0) contém uma data
                        col1_val = sheet_xlrd.cell_value(row_num, 0) if sheet_xlrd.ncols > 0 else ""
                        
                        # Verifica se é uma data (número de data do Excel ou string de data)
                        is_date = False
                        if isinstance(col1_val, (int, float)) and col1_val > 0:
                            # Pode ser uma data em formato numérico do Excel
                            try:
                                date_val = xlrd.xldate.xldate_as_datetime(col1_val, workbook_xlrd.datemode)
                                is_date = True
                                col1_val = date_val
                            except:
                                pass
                        elif isinstance(col1_val, str):
                            # Verifica se é string que parece data no formato ##/##/####
                            date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'
                            is_date = bool(re.match(date_pattern, col1_val))
                        
                        if is_date:
                            # Se encontrou data formatada, coleta dados das colunas 1, 2, 4
                            col2_val = sheet_xlrd.cell_value(row_num, 1) if sheet_xlrd.ncols > 1 else ""
                            col4_val = sheet_xlrd.cell_value(row_num, 3) if sheet_xlrd.ncols > 3 else ""
                            if col2_val == 'Pag. Boletos':
                                sair = True
                                break
                            x=col1_val,col2_val,col4_val
                            arq_data.append(x)
                            
                    except:
                        continue
                    if sair:
                        break
            
            else:  # Para arquivos .xlsx
                # Carrega o arquivo com openpyxl para verificar formatação
                workbook = load_workbook(caminho_arquivo, data_only=False)
                sheet = workbook.active
                
                st.write(f"Processando: {arquivo}")
                
                # Processa linha por linha
                for row_num in range(1, sheet.max_row + 1):
                    cell_a = sheet.cell(row=row_num, column=1)  # Coluna A (1)
                    
                    if is_date_formatted(cell_a):
                        # Se encontrou data formatada, coleta dados das colunas 1, 2, 4
                        col1_val = sheet.cell(row=row_num, column=1).value
                        col2_val = sheet.cell(row=row_num, column=2).value
                        col4_val = sheet.cell(row=row_num, column=4).value
                        
                        if col2_val == 'Pag. Boletos':
                            sair = True
                            break 
                        x=col1_val,col2_val,col4_val
                        arq_data.append(x)
                    elif row_num > 1 and arq_data:
                        # Se não é data e já coletou dados, verifica se deve parar
                        # Continua coletando até não haver mais datas consecutivas
                        pass
                
                workbook.close()
            
        except Exception as e:
            st.error(f"Erro ao processar {arquivo}: {str(e)}")
    
    return arq_data

def process_bradesco_files(arquivos, arq_data):
    """Processa arquivos Bradesco de 2025"""
      
    # Lista arquivos Bradesco de 2025
    arquivos_bradesco_2025 = []

    for arquivo in arquivos:
        
        if "2025" in arquivo and "Bradesco" in arquivo and (arquivo.endswith('.XLS') or arquivo.endswith('.XLSX') or arquivo.endswith('.xls') or arquivo.endswith('.xlsx')):
            arquivos_bradesco_2025.append(arquivo)
    
    for arquivo in arquivos_bradesco_2025:
        caminho_arquivo = arquivo
        
        try:
            # Para arquivos .xls antigos, usa xlrd
            if arquivo.endswith('.xls') or arquivo.endswith('.XLS'):
                workbook_xlrd = xlrd.open_workbook(caminho_arquivo)
                sheet_xlrd = workbook_xlrd.sheet_by_index(0)
                
                # Procura por "SALDO ANTERIOR"
                saldo_anterior_encontrado = False
                linha_inicio = 0
                total = False
                for row_num in range(sheet_xlrd.nrows):
                    for col_num in range(sheet_xlrd.ncols):                        
                        try:
                            cell_value = sheet_xlrd.cell_value(row_num, col_num)
                            if isinstance(cell_value, str) and "SALDO ANTERIOR" in cell_value.upper():
                                saldo_anterior_encontrado = True
                                linha_inicio = row_num + 1  # Próxima linha
                                break
                        except:
                            continue
                    if saldo_anterior_encontrado:
                        break
                
                if saldo_anterior_encontrado:                    
                    
                    # Coleta dados a partir da linha seguinte
                    for row_num in range(linha_inicio, sheet_xlrd.nrows):
                        try:
                            col1_val = sheet_xlrd.cell_value(row_num, 0) if sheet_xlrd.ncols > 0 else ""
                            col2_val = sheet_xlrd.cell_value(row_num, 1) if sheet_xlrd.ncols > 1 else ""
                            col3_val = sheet_xlrd.cell_value(row_num, 3) if sheet_xlrd.ncols > 3 else ""
                            col4_val = sheet_xlrd.cell_value(row_num, 4) if sheet_xlrd.ncols > 4 else ""
                            
                            # Se coluna 3 está vazia, usa coluna 4
                            terceira_coluna = col4_val if (not col3_val or col3_val == "") else col3_val
                            
                            if col1_val == "Total" or col2_val == None:
                                total = True
                                break
                            x=col1_val,col2_val,terceira_coluna                        
                            arq_data.append(x)                           
                    
                        except:
                            continue
                        if total:
                            break
                    
            else:  # Para arquivos .xlsx
                workbook = load_workbook(caminho_arquivo)
                sheet = workbook.active
                
                # Procura por "SALDO ANTERIOR"
                saldo_anterior_encontrado = False
                linha_inicio = 0
                total = False
                for row_num in range(1, sheet.max_row + 1):
                    for col_num in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_num, column=col_num).value
                        if cell_value and isinstance(cell_value, str) and "SALDO ANTERIOR" in cell_value.upper():
                            saldo_anterior_encontrado = True
                            linha_inicio = row_num + 1  # Próxima linha
                            break
                    if saldo_anterior_encontrado:
                        break
                
                if saldo_anterior_encontrado:                    
                    
                    # Coleta dados a partir da linha seguinte
                    for row_num in range(linha_inicio, sheet.max_row + 1):
                        col1_val = sheet.cell(row=row_num, column=1).value
                        col2_val = sheet.cell(row=row_num, column=2).value
                        col3_val = sheet.cell(row=row_num, column=4).value
                        col4_val = sheet.cell(row=row_num, column=5).value
                       
                        # Se coluna 3 está vazia, usa coluna 4
                        terceira_coluna = col4_val if (not col3_val or col3_val == "") else col3_val
                        if col1_val == "Total" or col2_val == None:
                            total = True
                        x=col1_val,col2_val,terceira_coluna   
                        if total:
                            break                     
                        arq_data.append(x)
                        
                workbook.close()
        
        except Exception as e:
            st.error(f"Erro ao processar {arquivo}: {str(e)}")
        
    return arq_data



def descricao(df_bradesco):
       
    # Processar e alterar os itens diretamente na lista df_bradesco
    df_bradesco_atualizado = []
    
    for i in df_bradesco:
        if len(i) > 1:  # Verifica se a tupla tem pelo menos 2 elementos
            # Pega o item 1 e converte para string
            item_original = str(i[1]) if i[1] is not None else ""
            
            # Divide o texto por espaços
            partes = item_original.split()
            
            # Filtra as partes, removendo apenas as que contêm SOMENTE números
            partes_filtradas = []
            # Lista de palavras a serem removidas
            palavras_remover = ['PAGAMENTO', 'PIX', 'ELETRON', 'COBRANCA', 'PAGTO', 'ELETRÔNICO', 'RECEBIMENTO', 
                                'TRANSF', 'TED', 'TRANSFERÊNCIA','BOLETO DE LIQUIDAÇÃO', 'LIQUIDAÇÃO',
                                'BOLETO', 'LIQUIDACAO','SICREDI ']
            
            for parte in partes:
                # Remove caracteres especiais da parte para análise
                parte_limpa = re.sub(r'[^A-Za-zÀ-ÿ0-9]', '', parte)
                parte_upper = parte_limpa.upper()
                
                # Verifica se a parte não é uma das palavras a serem removidas
                if parte_upper not in palavras_remover:
                    # Se a parte limpa não é apenas números OU se contém letras, mantém
                    if not parte_limpa.isdigit() or any(c.isalpha() for c in parte_limpa):
                        # Remove caracteres especiais mas mantém a parte
                        parte_sem_especiais = re.sub(r'[^A-Za-zÀ-ÿ0-9]', '', parte)
                        if parte_sem_especiais:  # Só adiciona se não ficou vazia
                            partes_filtradas.append(parte_sem_especiais)
            
            # Reconstrói o texto com espaços entre as posições mantidas
            item_limpo = ' '.join(partes_filtradas)
            
            # Converte para maiúsculas
            item_limpo = item_limpo.upper()
            
            # Cria nova tupla com o item 1 tratado
            if len(i) >= 3:
                nova_tupla = (i[0], item_limpo if item_limpo else i[1], i[2])
            else:
                nova_tupla = (i[0], item_limpo if item_limpo else i[1])
            
            # Aplicar padrões de descrição
            if len(nova_tupla) > 1 and nova_tupla[1]:
                descricao_padronizada = str(nova_tupla[1])
                
                # BARRIER para BARRIER TERCEIRIZACAO
                if "BARRIER" in descricao_padronizada.upper():
                    descricao_padronizada = "BARRIER TERCEIRIZACAO"
                # AURIGA para AURIGA FUNDO DE INVESTIMENTO
                elif "AURIGA" in descricao_padronizada.upper():
                    descricao_padronizada = "AURIGA FUNDO DE INVESTIMENTO"
                # BB CLAIM para BB CLAIM FUNDO DE INVESTIMENTO
                elif "BB CLAIM" in descricao_padronizada.upper():
                    descricao_padronizada = "BB CLAIM FUNDO DE INVESTIMENTO"
                # EVOLUCAO para EVOLUCAO AUDITORES
                elif "EVOLUCAO" in descricao_padronizada.upper():
                    descricao_padronizada = "EVOLUCAO AUDITORES"
                # EXCELSIOR para EXCELSIOR FUNDO DE INVESTIMENTO
                elif "EXCELSIOR" in descricao_padronizada.upper():
                    descricao_padronizada = "EXCELSIOR FUNDO DE INVESTIMENTO"
                # JL MORAIS para JL MORAIS SOLUCOES
                elif "MORAIS" in descricao_padronizada.upper():
                    descricao_padronizada = "JL MORAIS SOLUCOES"
                # SBC OPORTUNIDADE para SBC OPORTUNIDADE FUNDO DE
                elif "SBC OPORTUNIDADE" in descricao_padronizada.upper():
                    descricao_padronizada = "SBC OPORTUNIDADE FUNDO DE"
                # TEDTRANSF ELET DISPON REMETNDMP para NDMP I FIDC
                elif "TEDTRANSF ELET DISPON REMETNDMP" in descricao_padronizada.upper():
                    descricao_padronizada = "NDMP I FIDC"
                # TRANSFERÊNCIA DE 3I HOLDING LTDA para 3I HOLDING LTDA
                elif "3I" in descricao_padronizada.upper():
                    descricao_padronizada = "3I HOLDING LTDA"
                # TRANSFERÊNCIA DES 4I CAPITAL LTDA para 4I CAPITAL LTDA
                elif "4I CAPITAL" in descricao_padronizada.upper():
                    descricao_padronizada = "4I CAPITAL LTDA"
                # TRANSFERÊNCIA DE IGOR JEFFERSON LIMA C para IGOR JEFFERSON LIMA C
                elif "IGOR JEFFERSON" in descricao_padronizada.upper():
                    descricao_padronizada = "IGOR JEFFERSON LIMA C"
                # PROCESSO ANBIMA para ANBIMA ASSOC BR
                elif "ANBIMA" in descricao_padronizada.upper():
                    descricao_padronizada = "ANBIMA ASSOC BR"
                # NIO DIGITAL para NDMP I FIDC
                elif "NIO DIGITAL" in descricao_padronizada.upper():
                    descricao_padronizada = "NDMP I FIDC"
                # VANEY para VANEY IORI
                elif "VANEY" in descricao_padronizada.upper():
                    descricao_padronizada = "VANEY IORI"
                # LOCALIZA para LOCALIZA FLEET S A
                elif "FLEET" in descricao_padronizada.upper():
                    descricao_padronizada = "LOCALIZA FLEET S A"
                # PREFEITURA para PREFEITURA MUNI
                elif "PREFEITURA" in descricao_padronizada.upper():
                    descricao_padronizada = "PREFEITURA MUNICIPAL"
                # TRIAGEM para TRIAGEM CONSULTORIA
                elif "TRIAGEM" in descricao_padronizada.upper():
                    descricao_padronizada = "TRIAGEM CONSULTORIA"
                # V IORI para V IORI ADVISORY
                elif "V IORI" in descricao_padronizada.upper():
                    descricao_padronizada = "V IORI ADVISORY"
                # RENTABINVEST para INVEST FACEL
                elif "RENTABINVEST" in descricao_padronizada.upper():
                    descricao_padronizada = "INVEST FACIL"
                # LIF DESENVO para LIF DESENVOLVIMENTO
                elif "LIF DESENVO" in descricao_padronizada.upper():
                    descricao_padronizada = "LIF DESENVOLVIMENTO"
                # LUIGINO para LUIGINO IORI FILHO
                elif "LUIGINO" in descricao_padronizada.upper():
                    descricao_padronizada = "LUIGINO IORI FILHO"
                # TARIFA para TARIFA BANCARIA
                elif "TARIFA" in descricao_padronizada.upper():
                    descricao_padronizada = "TARIFA BANCARIA"
                # OPERACAO CAPITAL GIRO para OPERACAO CAPITAL GIRO
                elif "CAPITAL GIRO" in descricao_padronizada.upper():
                    descricao_padronizada = "OPERACAO CAPITAL GIRO"
                elif "CEF MATRIZ" in descricao_padronizada.upper():
                    descricao_padronizada = "OCAIXA ECONOMICA FED"
                
                # Atualizar tupla com descrição padronizada
                if descricao_padronizada != str(nova_tupla[1]):
                    if len(nova_tupla) >= 3:
                        nova_tupla = (nova_tupla[0], descricao_padronizada, nova_tupla[2])
                    else:
                        nova_tupla = (nova_tupla[0], descricao_padronizada)
            
            df_bradesco_atualizado.append(nova_tupla)
        else:
            # Se a tupla não tem item 1, mantém como está
            df_bradesco_atualizado.append(i)
        
    return df_bradesco_atualizado


def carregar_valores_manuais():
    """Carrega valores manuais salvos do arquivo JSON"""
    arquivo_valores = 'valores_manuais_projecao.json'
    if os.path.exists(arquivo_valores):
        try:
            with open(arquivo_valores, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def salvar_valores_manuais(valores_manuais):
    """Salva valores manuais no arquivo JSON"""
    arquivo_valores = 'valores_manuais_projecao.json'
    try:
        with open(arquivo_valores, 'w', encoding='utf-8') as f:
            json.dump(valores_manuais, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False

def carregar_classificacoes():
    """Carrega classificações existentes do arquivo JSON"""
    arquivo_classificacoes = "classificacoes_descricoes.json"
    if os.path.exists(arquivo_classificacoes):
        try:
            with open(arquivo_classificacoes, 'r', encoding='utf-8') as f:
                dados = json.load(f)
                
                # Migrar formato antigo para novo formato se necessário
                dados_migrados = {}
                migrou = False
                
                for descricao, info in dados.items():
                    if isinstance(info, str):
                        # Formato antigo: só classificação
                        dados_migrados[descricao] = {
                            "classificacao": info,
                            "recorrencia": None  # Será definida pelo usuário
                        }
                        migrou = True
                    elif isinstance(info, dict):
                        # Formato novo: já tem classificação e recorrência
                        dados_migrados[descricao] = info
                    else:
                        # Formato inválido, criar novo
                        dados_migrados[descricao] = {
                            "classificacao": "NÃO CLASSIFICADO",
                            "recorrencia": None
                        }
                        migrou = True
                
                # Se houve migração, salvar o novo formato
                if migrou:
                    salvar_classificacoes(dados_migrados)
                
                return dados_migrados
        except:
            return {}
    return {}


def salvar_classificacoes(classificacoes):
    """Salva classificações no arquivo JSON"""
    arquivo_classificacoes = "classificacoes_descricoes.json"
    try:
        with open(arquivo_classificacoes, 'w', encoding='utf-8') as f:
            json.dump(classificacoes, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


def verificar_classificacoes_sem_recorrencia():
    """Verifica se existem classificações sem recorrência definida"""
    classificacoes = carregar_classificacoes()
    sem_recorrencia = []
    
    for descricao, info in classificacoes.items():
        if isinstance(info, dict) and info.get('recorrencia') is None:
            sem_recorrencia.append(descricao)
    
    return sem_recorrencia


def obter_descricoes_unicas(dados_completos):
    """Obtém lista de descrições únicas dos dados"""
    descricoes = set()
    for registro in dados_completos:
        if len(registro) >= 2 and registro[1]:
            descricoes.add(str(registro[1]).strip())
    return sorted(list(descricoes))


def formulario_classificacao(dados_completos):
    """Cria formulário para classificação das descrições"""
    st.markdown("---")
    st.subheader("📝 Classificação de Descrições")
    
    # Opções de classificação disponíveis
    opcoes_classificacao = [
        "",
        "RECEITAS",
        "EMPRÉSTIMOS",
        "REEMBOLSO", 
        "CONTA CORRENTE",
        "APLICAÇÃO FINANCEIRA",
        "TRANSFERENCIA ENTRE CONTAS",
        "IMPOSTOS",
        "FOLHA CLT",
        "FOLHA PJ",
        "ENCARGOS",
        "ADMINISTRATIVA",
        "ASSESSORIA JURIDICA",
        "ASSESSORIA CONTABIL",
        "DESPESAS FINANCEIRAS",
        "DESPESAS COMERCIAIS",
        "SOFTWARE",
        "PMT EMPRESTIMOS",
        "INVESTIMENTOS",
        "DESPESAS IMÓVEL",        
        "ADIANTAMENTO A FORNECEDORES"
    ]
    
    # Opções de recorrência
    opcoes_recorrencia = ["", "RE", "N_RE"]
    
    # Carregar classificações existentes
    classificacoes_existentes = carregar_classificacoes()
    
    # Verificar classificações sem recorrência definida
    classificacoes_sem_recorrencia = verificar_classificacoes_sem_recorrencia()
    
    # Mostrar alerta se existem classificações sem recorrência
    if classificacoes_sem_recorrencia:
        st.warning(f"⚠️ **{len(classificacoes_sem_recorrencia)} classificações precisam ter a recorrência definida (RE/N_RE)!**")
        
        # Formulário para definir recorrências em lote
        st.subheader("🔄 Definir Recorrência das Classificações Existentes")
        
        with st.expander("Clique aqui para definir as recorrências", expanded=True):
            with st.form("recorrencia_form"):
                st.write("**Defina se cada classificação é Recorrente (RE) ou Não Recorrente (N_RE):**")
                
                recorrencias_update = {}
                
                # Dividir em colunas para melhor layout
                num_cols = 2
                cols = st.columns(num_cols)
                
                for i, descricao in enumerate(classificacoes_sem_recorrencia):
                    col = cols[i % num_cols]
                    
                    with col:
                        info_atual = classificacoes_existentes[descricao]
                        classificacao_atual = info_atual.get('classificacao', 'NÃO CLASSIFICADO') if isinstance(info_atual, dict) else info_atual
                        
                        st.write(f"**{descricao}**")
                        st.write(f"*Classificação: {classificacao_atual}*")
                        
                        recorrencia = st.selectbox(
                            "Recorrência:",
                            ["Selecione...", "RE (Recorrente)", "N_RE (Não Recorrente)"],
                            key=f"rec_{i}"
                        )
                        
                        if recorrencia != "Selecione...":
                            valor_recorrencia = "RE" if recorrencia.startswith("RE") else "N_RE"
                            recorrencias_update[descricao] = valor_recorrencia
                        
                        st.markdown("---")
                
                if st.form_submit_button("💾 Salvar Recorrências"):
                    if recorrencias_update:
                        # Atualizar classificações com recorrências
                        for desc, rec in recorrencias_update.items():
                            if desc in classificacoes_existentes:
                                if isinstance(classificacoes_existentes[desc], str):
                                    # Converter formato antigo para novo
                                    classificacoes_existentes[desc] = {
                                        "classificacao": classificacoes_existentes[desc],
                                        "recorrencia": rec
                                    }
                                else:
                                    # Atualizar formato novo
                                    classificacoes_existentes[desc]["recorrencia"] = rec
                        
                        if salvar_classificacoes(classificacoes_existentes):
                            st.success(f"✅ {len(recorrencias_update)} recorrências salvas com sucesso!")
                            st.rerun()
                        else:
                            st.error("❌ Erro ao salvar recorrências!")
                    else:
                        st.warning("⚠️ Selecione pelo menos uma recorrência para salvar.")
    
    # Obter descrições únicas
    descricoes_unicas = obter_descricoes_unicas(dados_completos)
    
    # Filtrar descrições não classificadas
    descricoes_nao_classificadas = [desc for desc in descricoes_unicas 
                                    if desc not in classificacoes_existentes]
    
    # Adicionar seletor para editar classificações existentes
    st.subheader("✏️ Editar Classificações Existentes")
    
    if classificacoes_existentes:
        # Selectbox para escolher descrição para editar
        descricoes_para_editar = ["Selecione uma descrição para editar..."] + sorted(list(classificacoes_existentes.keys()))
        descricao_selecionada = st.selectbox(
            "Escolha uma descrição para editar:",
            descricoes_para_editar,
            key="edit_selector"
        )
        
        if descricao_selecionada != "Selecione uma descrição para editar...":
            info_atual = classificacoes_existentes[descricao_selecionada]
            
            # Extrair classificação e recorrência atuais
            if isinstance(info_atual, dict):
                classificacao_atual = info_atual.get('classificacao', '')
                recorrencia_atual = info_atual.get('recorrencia', '')
            else:
                # Formato antigo - só classificação
                classificacao_atual = info_atual
                recorrencia_atual = ''
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.write(f"**Descrição:** {descricao_selecionada}")
                st.write(f"**Classificação atual:** {classificacao_atual}")
                if recorrencia_atual:
                    recorrencia_desc = "Recorrente" if recorrencia_atual == "RE" else "Não Recorrente"
                    st.write(f"**Recorrência atual:** {recorrencia_atual} ({recorrencia_desc})")
                else:
                    st.write(f"**Recorrência atual:** *Não definida*")
                
                # Dropdown para nova classificação
                try:
                    index_atual = opcoes_classificacao.index(classificacao_atual)
                except ValueError:
                    index_atual = 0
                
                nova_classificacao = st.selectbox(
                    "Nova classificação:",
                    opcoes_classificacao,
                    index=index_atual,
                    key="new_classification"
                )
                
                # Dropdown para nova recorrência
                try:
                    if recorrencia_atual:
                        index_rec = opcoes_recorrencia.index(recorrencia_atual)
                    else:
                        index_rec = 0
                except ValueError:
                    index_rec = 0
                
                nova_recorrencia = st.selectbox(
                    "Recorrência:",
                    ["", "RE (Recorrente)", "N_RE (Não Recorrente)"],
                    index=index_rec if recorrencia_atual else 0,
                    key="new_recurrence"
                )
            
            with col2:
                st.write("")
                st.write("")
                
                if st.button("💾 Atualizar Classificação", key="update_btn"):
                    # Extrair valor da recorrência selecionada
                    valor_recorrencia = ""
                    if nova_recorrencia.startswith("RE"):
                        valor_recorrencia = "RE"
                    elif nova_recorrencia.startswith("N_RE"):
                        valor_recorrencia = "N_RE"
                    
                    # Verificar se houve alteração
                    alterou_classificacao = nova_classificacao != classificacao_atual
                    alterou_recorrencia = valor_recorrencia != recorrencia_atual
                    
                    if alterou_classificacao or alterou_recorrencia:
                        # Criar nova estrutura de dados
                        classificacoes_existentes[descricao_selecionada] = {
                            "classificacao": nova_classificacao,
                            "recorrencia": valor_recorrencia if valor_recorrencia else None
                        }
                        
                        if salvar_classificacoes(classificacoes_existentes):
                            msg_sucesso = "✅ Atualizado:"
                            if alterou_classificacao:
                                msg_sucesso += f" Classificação: **{nova_classificacao}**"
                            if alterou_recorrencia:
                                rec_desc = "Recorrente" if valor_recorrencia == "RE" else "Não Recorrente" if valor_recorrencia == "N_RE" else "Não definida"
                                msg_sucesso += f" Recorrência: **{rec_desc}**"
                            st.success(msg_sucesso)
                            st.rerun()
                        else:
                            st.error("❌ Erro ao salvar a atualização!")
                    else:
                        st.info("ℹ️ Nenhuma alteração detectada.")
                
                if st.button("🗑️ Excluir Classificação", key="delete_btn"):
                    del classificacoes_existentes[descricao_selecionada]
                    if salvar_classificacoes(classificacoes_existentes):
                        st.success("✅ Classificação excluída!")
                        st.rerun()
                    else:
                        st.error("❌ Erro ao excluir!")
    
    st.markdown("---")
    
    if not descricoes_nao_classificadas:
        st.success("✅ Todas as descrições já estão classificadas!")
        
        # Mostrar classificações existentes
        if st.checkbox("Mostrar todas as classificações cadastradas"):
            st.write("**Classificações cadastradas:**")
            for desc, info in sorted(classificacoes_existentes.items()):
                if isinstance(info, dict):
                    classificacao = info.get('classificacao', 'NÃO CLASSIFICADO')
                    recorrencia = info.get('recorrencia', 'Não definida')
                    rec_desc = ""
                    if recorrencia == "RE":
                        rec_desc = " (Recorrente)"
                    elif recorrencia == "N_RE":
                        rec_desc = " (Não Recorrente)"
                    elif recorrencia is None or recorrencia == "Não definida":
                        rec_desc = " ⚠️ (Recorrência não definida)"
                    st.write(f"• {desc} → **{classificacao}**{rec_desc}")
                else:
                    # Formato antigo
                    st.write(f"• {desc} → **{info}** ⚠️ (Recorrência não definida)")
        return
    
    st.subheader("➕ Classificar Novas Descrições")
    st.write(f"**{len(descricoes_nao_classificadas)}** descrições precisam ser classificadas:")
    
    # Formulário para classificar
    with st.form("classificacao_form"):
        classificacoes_novas = {}
        
        # Dividir em colunas para melhor layout
        num_cols = 2
        cols = st.columns(num_cols)
        
        for i, descricao in enumerate(descricoes_nao_classificadas[:10]):  # Limitar a 10 por vez
            col = cols[i % num_cols]
            
            with col:
                st.write(f"**Descrição:** {descricao}")
                
                classificacao = st.selectbox(
                    "Classificação:",
                    ["Selecione..."] + opcoes_classificacao,
                    key=f"class_{i}",
                    index=0
                )
                
                recorrencia = st.selectbox(
                    "Recorrência:",
                    ["Selecione...", "RE (Recorrente)", "N_RE (Não Recorrente)"],
                    key=f"rec_{i}",
                    index=0
                )
                
                if classificacao != "Selecione..." and recorrencia != "Selecione...":
                    valor_recorrencia = "RE" if recorrencia.startswith("RE") else "N_RE"
                    classificacoes_novas[descricao] = {
                        "classificacao": classificacao,
                        "recorrencia": valor_recorrencia
                    }
                
                st.write("---")
        
        # Botões do formulário
        col1, col2 = st.columns([1, 1])
        
        with col1:
            submitted = st.form_submit_button("💾 Salvar Classificações")
        
        with col2:
            if len(descricoes_nao_classificadas) > 10:
                st.write(f"Restam {len(descricoes_nao_classificadas) - 10} descrições")
    
    # Processar envio do formulário
    if submitted and classificacoes_novas:
        # Mesclar com classificações existentes
        classificacoes_existentes.update(classificacoes_novas)
        
        # Salvar no arquivo
        if salvar_classificacoes(classificacoes_existentes):
            st.success(f"✅ {len(classificacoes_novas)} classificações salvas com sucesso!")
            st.rerun()
        else:
            st.error("❌ Erro ao salvar classificações!")


def aplicar_classificacoes(dados_completos):
    """Aplica classificações aos dados e retorna dados classificados"""
    classificacoes = carregar_classificacoes()
    dados_classificados = []
    
    for registro in dados_completos:
        if len(registro) >= 3:
            data, descricao, valor = registro[0], registro[1], registro[2]
            
            # Obter informação da classificação
            info_classificacao = classificacoes.get(str(descricao).strip(), "NÃO CLASSIFICADO")
            
            # Extrair apenas a classificação (formato novo ou antigo)
            if isinstance(info_classificacao, dict):
                classificacao = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                # Formato antigo - já é a classificação
                classificacao = info_classificacao
            
            # Adiciona classificação como quarta coluna
            dados_classificados.append((data, descricao, valor, classificacao))
        else:
            dados_classificados.append(registro + ("NÃO CLASSIFICADO",))
    
    return dados_classificados


def calcular_medias_recorrentes(dados_completos):
    """Calcula médias mensais apenas para itens com recorrencia='RE'"""
    classificacoes = carregar_classificacoes()
    
    # Dicionário para armazenar os dados apenas de itens recorrentes
    dados_recorrentes = {}
    
    # Processar cada registro
    for registro in dados_completos:
        if len(registro) >= 3:
            data, descricao, valor = registro[0], registro[1], registro[2]
            
            # Verificar se é recorrente
            info_classificacao = classificacoes.get(str(descricao).strip())
            if info_classificacao and isinstance(info_classificacao, dict):
                recorrencia = info_classificacao.get('recorrencia')
                if recorrencia == 'RE':  # Apenas itens recorrentes
                    
                    # Extrair mês da data
                    try:
                        if isinstance(data, datetime):
                            mes = data.month
                        elif isinstance(data, str):
                            # Tentar diferentes formatos de data
                            for formato in ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']:
                                try:
                                    data_obj = datetime.strptime(data, formato)
                                    mes = data_obj.month
                                    break
                                except:
                                    continue
                            else:
                                mes = 1  # Default para janeiro se não conseguir parsear
                        else:
                            mes = 1  # Default
                        
                        # Converter valor para float
                        if isinstance(valor, str):                
                            # Remove caracteres não numéricos exceto vírgula, ponto e sinal negativo
                            valor_limpo = re.sub(r'[^\d,.\-]', '', str(valor))
                            valor_limpo = valor_limpo.replace(',', '.')
                            
                            try:
                                valor_float = float(valor_limpo)
                            except:
                                valor_float = 0.0
                        else:
                            valor_float = float(valor) if valor else 0.0
                        
                        # Inicializar estrutura se necessário
                        if descricao not in dados_recorrentes:                    
                            dados_recorrentes[descricao] = {i: [] for i in range(1, 13)}  # Lista para cada mês
                        
                        # Adicionar valor à lista do mês correspondente                
                        dados_recorrentes[descricao][mes].append(valor_float)
                        
                    except Exception as e:
                        print(f"Erro processando registro recorrente: {e}")
                        continue
    
    # Calcular médias para cada descrição e mês
    medias_recorrentes = {}
    for descricao, meses_data in dados_recorrentes.items():
        medias_recorrentes[descricao] = {}
        for mes in range(1, 13):
            valores = meses_data[mes]
            if valores:
                # Calcular média dos valores do mês
                media = sum(valores) / len(valores)
                medias_recorrentes[descricao][mes] = media
            else:
                # Se não há dados para o mês, usar 0
                medias_recorrentes[descricao][mes] = 0.0
    
    return medias_recorrentes


def criar_tabela_por_classificacao(dados_classificados):
    """Cria tabela resumo por classificação"""
    if not dados_classificados:
        return ""
    
    # Organizar dados por classificação
    resumo_classificacao = {}
    
    for registro in dados_classificados:
        if len(registro) >= 4:
            data, descricao, valor, classificacao = registro[0], registro[1], registro[2], registro[3]
            
            # Converter valor para float
            if isinstance(valor, str):
                valor_limpo = re.sub(r'[^\d,.\-]', '', str(valor))
                valor_limpo = valor_limpo.replace(',', '.')
                try:
                    valor_float = float(valor_limpo)
                except:
                    valor_float = 0.0
            else:
                valor_float = float(valor) if valor else 0.0
            
            if classificacao not in resumo_classificacao:
                resumo_classificacao[classificacao] = 0.0
            
            resumo_classificacao[classificacao] += valor_float
    
    # Criar HTML da tabela de classificação
    html = """
    <div style="margin-top: 20px;">
    <h3>💼 Resumo por Classificação</h3>
    <table style="border-collapse: collapse; width: 100%; max-width: 800px;">
    <thead>
        <tr style="background-color: #f0f0f0;">
            <th style="border: 1px solid #ddd; padding: 12px; text-align: left;">Classificação</th>
            <th style="border: 1px solid #ddd; padding: 12px; text-align: right;">Total (R$)</th>
        </tr>
    </thead>
    <tbody>
    """
    
    # Ordenar por classificação
    for classificacao in sorted(resumo_classificacao.keys()):
        total = resumo_classificacao[classificacao]
        cor_texto = "color: #dc3545;" if total < 0 else "color: #28a745;"
        
        html += f"""
        <tr>
            <td style="border: 1px solid #ddd; padding: 8px;">{classificacao}</td>
            <td style="border: 1px solid #ddd; padding: 8px; text-align: right; {cor_texto}">
                R$ {total:,.2f}
            </td>
        </tr>
        """
    
    html += """
    </tbody>
    </table>
    </div>
    """
    
    return html


def criar_tabela_fluxo_futuro(dados_completos):
    """Cria tabela HTML de fluxo futuro baseada nas médias dos itens recorrentes - Seguindo estrutura da tabela mensal"""
    
    # Calcular médias dos itens recorrentes
    medias_recorrentes = calcular_medias_recorrentes(dados_completos)
    
    # Carregar valores manuais salvos
    valores_manuais = carregar_valores_manuais()
    
    if not medias_recorrentes:
        return """
        <div style="margin-top: 20px;">
        <h3>📈 Projeção de Fluxo Futuro - Próximos 12 Meses (Itens Recorrentes)</h3>
        <p style="color: #666; font-style: italic;">Nenhum item recorrente encontrado para projeção.</p>
        </div>
        """
    
    # Usar mesma estrutura da tabela mensal
    ordem_classificacoes = [
        "RECEITAS",
        "EMPRÉSTIMOS",
        "REEMBOLSO", 
        "CONTA CORRENTE",
        "APLICAÇÃO FINANCEIRA",
        "TRANSFERENCIA ENTRE CONTAS",
        "DESPESAS",
        "IMPOSTOS",
        "FOLHA CLT",
        "FOLHA PJ",
        "ENCARGOS",
        "ADMINISTRATIVA",
        "ASSESSORIA JURIDICA",
        "ASSESSORIA CONTABIL",
        "DESPESAS FINANCEIRAS",
        "DESPESAS COMERCIAIS",
        "SOFTWARE",
        "PMT EMPRESTIMOS",
        "INVESTIMENTOS",
        "DESPESAS IMÓVEL",        
        "ADIANTAMENTO A FORNECEDORES",
        "NÃO CLASSIFICADO"
    ]
    
    # Subcategorias de DESPESAS
    subcategorias_despesas = [
        "IMPOSTOS",
        "FOLHA CLT",
        "FOLHA PJ",
        "ENCARGOS",
        "ADMINISTRATIVA",
        "ASSESSORIA JURIDICA",
        "ASSESSORIA CONTABIL",
        "DESPESAS FINANCEIRAS",
        "DESPESAS COMERCIAIS",
        "SOFTWARE",
        "PMT EMPRESTIMOS",
        "INVESTIMENTOS",
        "DESPESAS IMÓVEL",
        "ADIANTAMENTO A FORNECEDORES"
    ]
    
    # Classificações de receitas
    classificacoes_receitas = [
        "RECEITAS",
        "EMPRÉSTIMOS",
        "REEMBOLSO",
        "CONTA CORRENTE",
        "APLICAÇÃO FINANCEIRA",
        "TRANSFERENCIA ENTRE CONTAS"
    ]
    
    # Carregar classificações
    classificacoes = carregar_classificacoes()
    
    # Nomes dos próximos 12 meses
    from datetime import datetime, timedelta
    import calendar
    
    meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                   'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    # Função para obter índice da classificação para ordenação (igual à tabela mensal)
    def obter_indice_classificacao_futuro(descricao):
        info_classificacao = classificacoes.get(str(descricao).strip(), "NÃO CLASSIFICADO")
        
        # Extrair apenas a classificação
        if isinstance(info_classificacao, dict):
            classificacao = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
        else:
            classificacao = info_classificacao
            
        try:
            return ordem_classificacoes.index(classificacao)
        except ValueError:
            return len(ordem_classificacoes)  # Colocar no final se não encontrar
    
    # Ordenar descrições por classificação (igual à tabela mensal)
    descricoes_ordenadas = sorted(medias_recorrentes.keys(), 
                                 key=lambda desc: (obter_indice_classificacao_futuro(desc), desc))
    
    # Calcular totais das receitas projetadas (igual à tabela mensal)
    def calcular_totais_receitas_futuras():
        totais_mes_receitas = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
        total_geral_receitas = 0.0
        
        for desc in medias_recorrentes.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            
            # Extrair apenas a classificação
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
                
            if classificacao_desc in classificacoes_receitas:
                # Verificar se há valor manual para esta descrição
                valor_manual = valores_manuais.get(str(desc), None)
                
                if valor_manual:
                    # Usar valor manual como valor mensal (não dividir por 12)
                    valor_mensal = float(valor_manual)
                    for mes in range(1, 13):
                        totais_mes_receitas[mes] += valor_mensal
                    total_geral_receitas += float(valor_manual) * 12
                else:
                    # Usar valores calculados
                    for mes in range(1, 13):
                        valor_medio = medias_recorrentes[desc].get(mes, 0.0)
                        totais_mes_receitas[mes] += valor_medio
                    total_geral_receitas += sum(medias_recorrentes[desc].values())
        
        return totais_mes_receitas, total_geral_receitas
    
    # Calcular totais das despesas projetadas (igual à tabela mensal)
    def calcular_totais_despesas_futuras():
        totais_mes_despesas = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
        total_geral_despesas = 0.0
        
        for desc in medias_recorrentes.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            
            # Extrair apenas a classificação
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
                
            if classificacao_desc in subcategorias_despesas:
                # Verificar se há valor manual para esta descrição
                valor_manual = valores_manuais.get(str(desc), None)
                
                if valor_manual:
                    # Usar valor manual como valor mensal (não dividir por 12)
                    valor_mensal = float(valor_manual)
                    for mes in range(1, 13):
                        totais_mes_despesas[mes] += valor_mensal
                    total_geral_despesas += float(valor_manual) * 12
                else:
                    # Usar valores calculados
                    for mes in range(1, 13):
                        valor_medio = medias_recorrentes[desc].get(mes, 0.0)
                        totais_mes_despesas[mes] += valor_medio
                    total_geral_despesas += sum(medias_recorrentes[desc].values())
        
        return totais_mes_despesas, total_geral_despesas
    
    # Calcular totais das receitas e despesas
    totais_mes_receitas, total_geral_receitas = calcular_totais_receitas_futuras()
    totais_mes_despesas, total_geral_despesas = calcular_totais_despesas_futuras()
    
    # CALCULAR SALDO INICIAL BASEADO NO ÚLTIMO SALDO DA TABELA MENSAL
    # Primeiro, calcular os dados da tabela mensal para obter o último saldo
    # Replicar a lógica da tabela mensal original
    tabela_dados = {}
    for registro in dados_completos:
        if len(registro) >= 3:
            data, descricao, valor = registro[0], registro[1], registro[2]
            
            try:
                if isinstance(data, datetime):
                    mes = data.month
                elif isinstance(data, str):
                    for formato in ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']:
                        try:
                            data_obj = datetime.strptime(data, formato)
                            mes = data_obj.month
                            break
                        except:
                            continue
                    else:
                        mes = 1
                else:
                    mes = 1
                
                if isinstance(valor, str):                
                    valor_limpo = re.sub(r'[^\d,.\-]', '', str(valor))
                    valor_limpo = valor_limpo.replace(',', '.')
                    try:
                        valor_float = float(valor_limpo)
                    except:
                        valor_float = 0.0
                else:
                    valor_float = float(valor) if valor else 0.0
                    
                if descricao not in tabela_dados:                    
                    tabela_dados[descricao] = {i: 0.0 for i in range(1, 13)}
                tabela_dados[descricao][mes] += valor_float
                
            except Exception as e:
                continue
    
    # Calcular totais da tabela mensal original para obter o último saldo
    def calcular_totais_receitas_original():
        totais_mes = {i: 0.0 for i in range(1, 13)}
        for desc in tabela_dados.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
            if classificacao_desc in classificacoes_receitas:
                dados_mes = tabela_dados[desc]
                for mes in range(1, 13):
                    totais_mes[mes] += dados_mes[mes]
        return totais_mes
    
    def calcular_totais_despesas_original():
        totais_mes = {i: 0.0 for i in range(1, 13)}
        for desc in tabela_dados.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
            if classificacao_desc in subcategorias_despesas:
                dados_mes = tabela_dados[desc]
                for mes in range(1, 13):
                    totais_mes[mes] += dados_mes[mes]
        return totais_mes
    
    # Calcular saldo inicial baseado no último mês da tabela original
    totais_receitas_original = calcular_totais_receitas_original()
    totais_despesas_original = calcular_totais_despesas_original()
    
    saldo_inicial_base = 272801.75  # Mesmo da tabela mensal
    saldo_acumulado = saldo_inicial_base
    
    # Calcular saldo até dezembro (último mês)
    for mes in range(1, 13):
        movimento_mes = totais_receitas_original[mes] + totais_despesas_original[mes]
        saldo_acumulado += movimento_mes
    
    # O saldo final de dezembro será o saldo inicial da projeção
    saldo_inicial_projecao = saldo_acumulado
    
    # Calcular saldo bancário cumulativo futuro
    saldo_acumulado_futuro = saldo_inicial_projecao
    saldos_futuros_cumulativos = {}
    
    # Calcular saldo acumulado mês a mês (projeção)
    for mes in range(1, 13):
        movimento_mes = totais_mes_receitas[mes] + totais_mes_despesas[mes]
        saldo_acumulado_futuro += movimento_mes
        saldos_futuros_cumulativos[mes] = saldo_acumulado_futuro
    
    # Criar HTML da tabela (seguindo estrutura da tabela mensal)
    html = """
    <div style="margin-top: 20px;">
    <div style="max-height: 600px; overflow-y: auto;">
    <table border="1" style="width:100%; border-collapse: collapse;">
    <tr style="background-color: #f0f0f0;">
        <th style="text-align: left; padding: 8px; background-color: #f0f0f0;">Descrição (Recorrente)</th>
        <th style="text-align: center; padding: 8px; background-color: #f0f0f0;">Total</th>
    """
    
    # Adicionar cabeçalhos dos meses (igual à tabela mensal)
    for mes_nome in meses_nomes:
        html += f'<th style="text-align: center; padding: 8px; background-color: #f0f0f0; font-size: 14px;">{mes_nome}</th>'
    
    html += '</tr>'
    
    # Adicionar linha de saldo bancário projetado (igual à tabela mensal)
    html += f'<tr style="background-color: rgba(255, 215, 0, 0.4); font-weight: bold; border: 3px solid #FFD700;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 20px; font-weight: bold; color: #4169E1;">💰 SALDO BANCÁRIO PROJETADO</td>'
    html += f'<td style="padding: 8px; text-align: center; font-size: 16px; font-weight: bold; color: #666;">Saldo Inicial: {int(saldo_inicial_projecao):,}</td>'
    
    # Adicionar saldo projetado de cada mês
    for mes in range(1, 13):
        saldo_mes_projetado = saldos_futuros_cumulativos[mes]
        cor_saldo_mes = '#0066CC' if saldo_mes_projetado >= 0 else '#DC143C'
        html += f'<td style="padding: 8px; color: {cor_saldo_mes}; text-align: center; font-weight: bold; font-size: 18px;">{int(saldo_mes_projetado):,}</td>'
    
    html += '</tr>'
    
    # Linha de divisão
    html += f'<tr style="height: 3px; border: none;"><td colspan="14" style="background-color: #666; height: 3px; border: none; padding: 0;"></td></tr>'
    
    # Adicionar linha RECEITA / DESPESAS projetadas
    total_receita_despesas_futuras = total_geral_receitas + total_geral_despesas
    html += f'<tr style="background-color: rgba(255, 165, 0, 0.4); font-weight: bold; border: 2px solid #FFA500;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold; color: #FF4500;">💰 RECEITA / DESPESAS PROJETADAS</td>'
    
    cor_total_receita_despesas = '#228B22' if total_receita_despesas_futuras >= 0 else '#DC143C'
    html += f'<td style="padding: 8px; color: {cor_total_receita_despesas}; text-align: center; font-size: 18px; font-weight: bold;">{int(total_receita_despesas_futuras):,}</td>'
    
    # Adicionar valores mensais projetados (receitas + despesas por mês)
    for mes in range(1, 13):
        valor_mes_total = totais_mes_receitas[mes] + totais_mes_despesas[mes]
        if valor_mes_total != 0:
            cor_valor = '#228B22' if valor_mes_total >= 0 else '#DC143C'
            html += f'<td style="padding: 8px; color: {cor_valor}; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes_total):,}</td>'
        else:
            html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px; color: #FF4500;">-</td>'
    
    html += '</tr>'
    
    # Linha de divisão
    html += f'<tr style="height: 3px; border: none;"><td colspan="14" style="background-color: #666; height: 3px; border: none; padding: 0;"></td></tr>'
    
    # Adicionar linha de total das receitas projetadas
    html += f'<tr style="background-color: rgba(144, 238, 144, 0.6); font-weight: bold; border: 2px solid #90EE90;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold; color: #000080;">💰 RECEITAS/APLICAÇÕES/EMPRESTIMOS PROJETADAS</td>'
    html += f'<td style="padding: 8px; color: #000080; text-align: center; font-size: 18px; font-weight: bold;">{int(total_geral_receitas):,}</td>'
    
    # Adicionar totais de cada mês para receitas
    for mes in range(1, 13):
        valor_mes = totais_mes_receitas[mes]
        if valor_mes != 0:
            html += f'<td style="padding: 8px; color: #000080; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes):,}</td>'
        else:
            html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px; color: #000080;">-</td>'
    
    html += '</tr>'
    
    # Linha de divisão mais grossa
    html += f'<tr style="height: 8px; border: none;"><td colspan="14" style="background-color: #00CED1; height: 8px; border: none; padding: 0; border-top: 3px solid #008B8B; border-bottom: 2px solid #008B8B;"></td></tr>'
    
    # Função para calcular totais por classificação (igual à tabela mensal)
    def calcular_totais_classificacao_futura(classificacao):
        totais_mes = {i: 0.0 for i in range(1, 13)}
        total_geral = 0.0
        
        if classificacao == "DESPESAS":
            for desc in descricoes_ordenadas:
                info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
                if isinstance(info_classificacao, dict):
                    classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
                else:
                    classificacao_desc = info_classificacao
                if classificacao_desc in subcategorias_despesas:
                    # Verificar se há valor manual
                    valor_manual = valores_manuais.get(str(desc), None)
                    if valor_manual:
                        valor_mensal = float(valor_manual)
                        for mes in range(1, 13):
                            totais_mes[mes] += valor_mensal
                        total_geral += float(valor_manual) * 12
                    else:
                        for mes in range(1, 13):
                            totais_mes[mes] += medias_recorrentes[desc].get(mes, 0.0)
                        total_geral += sum(medias_recorrentes[desc].values())
        else:
            for desc in descricoes_ordenadas:
                info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
                if isinstance(info_classificacao, dict):
                    classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
                else:
                    classificacao_desc = info_classificacao
                if classificacao_desc == classificacao:
                    # Verificar se há valor manual
                    valor_manual = valores_manuais.get(str(desc), None)
                    if valor_manual:
                        valor_mensal = float(valor_manual)
                        for mes in range(1, 13):
                            totais_mes[mes] += valor_mensal
                        total_geral += float(valor_manual) * 12
                    else:
                        for mes in range(1, 13):
                            totais_mes[mes] += medias_recorrentes[desc].get(mes, 0.0)
                        total_geral += sum(medias_recorrentes[desc].values())
        
        return totais_mes, total_geral
    
    # Adicionar dados de cada descrição ordenada com separadores por classificação específica
    classificacao_anterior = None
    despesas_ja_adicionada = False
    
    for descricao in descricoes_ordenadas:
        # Obter classificação atual
        info_classificacao = classificacoes.get(str(descricao).strip(), "NÃO CLASSIFICADO")
        
        if isinstance(info_classificacao, dict):
            classificacao_atual = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
        else:
            classificacao_atual = info_classificacao
        
        # Se é uma subcategoria de DESPESAS e ainda não foi adicionada a linha DESPESAS geral
        if classificacao_atual in subcategorias_despesas and not despesas_ja_adicionada:
            # Adicionar linha de total DESPESAS geral
            totais_despesas_mes, total_despesas_geral = calcular_totais_classificacao_futura("DESPESAS")
            
            html += f'<tr style="background-color: rgba(255, 182, 193, 0.6); font-weight: bold; border: 2px solid #FFB6C1;">'
            html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold; color: #8B0000;">💸 DESPESAS PROJETADAS</td>'
            html += f'<td style="padding: 8px; color: #8B0000; text-align: center; font-size: 18px; font-weight: bold;">{int(total_despesas_geral):,}</td>'
            
            for mes in range(1, 13):
                valor_mes = totais_despesas_mes[mes]
                if valor_mes != 0:
                    html += f'<td style="padding: 8px; color: #8B0000; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes):,}</td>'
                else:
                    html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px; color: #8B0000;">-</td>'
            
            html += '</tr>'
            despesas_ja_adicionada = True
        
        # Adicionar linha de separação apenas para classificações específicas dentro das despesas
        if classificacao_atual != classificacao_anterior:
            
            # Adicionar separador para a nova classificação (se for subcategoria de despesas)
            if classificacao_atual in subcategorias_despesas:
                # Calcular totais para esta classificação específica de despesa
                totais_class_mes, total_class_geral = calcular_totais_classificacao_futura(classificacao_atual)
                
                # Só adicionar se houver itens recorrentes nesta classificação
                if total_class_geral != 0:
                    html += f'<tr style="background-color: rgba(255, 192, 203, 0.4); font-weight: bold; border: 1px solid #FFB6C1;">'
                    html += f'<td style="padding: 8px; text-align: center; font-size: 16px; font-weight: bold; color: #B22222;">├─ {classificacao_atual}</td>'
                    html += f'<td style="padding: 8px; color: #B22222; text-align: center; font-size: 16px; font-weight: bold;">{int(total_class_geral):,}</td>'
                    
                    for mes in range(1, 13):
                        valor_mes = totais_class_mes[mes]
                        if valor_mes != 0:
                            html += f'<td style="padding: 8px; color: #B22222; text-align: center; font-weight: bold; font-size: 16px;">{int(valor_mes):,}</td>'
                        else:
                            html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 16px; color: #B22222;">-</td>'
                    
                    html += '</tr>'
        
        # Adicionar linha da descrição individual
        total_descricao = sum(medias_recorrentes[descricao].values())
        cor_total = '#228B22' if total_descricao >= 0 else '#DC143C'
        
        # Cor de fundo baseada na classificação
        if classificacao_atual in classificacoes_receitas:
            cor_bg_desc = "#f0fff0"  # Verde claro para receitas
            prefixo = "   "
        elif classificacao_atual in subcategorias_despesas:
            cor_bg_desc = "#fff5f5"  # Rosa bem claro para despesas
            prefixo = "     • "  # Maior indentação para itens de despesa
        else:
            cor_bg_desc = "#f9f9f9"  # Cinza claro
            prefixo = "   "
        
        # Criar ID único para o input (usando descricao sem caracteres especiais)
        input_id = f"manual_{abs(hash(str(descricao)))}"
        
        # Verificar se há valor manual salvo para esta descrição
        valor_manual_salvo = valores_manuais.get(str(descricao), "")
        
        # Se há valor manual, usar ele para calcular o total e valores mensais
        if valor_manual_salvo:
            total_exibido = float(valor_manual_salvo) * 12  # Multiplicar por 12 para mostrar total anual
            cor_total = '#228B22' if total_exibido >= 0 else '#DC143C'
        else:
            total_exibido = total_descricao
        
        # Adicionar indicador visual se o valor é manual
        indicador_manual = " 🔧" if valor_manual_salvo else ""
        
        html += f'<tr style="background-color: {cor_bg_desc};" id="row_{input_id}">'
        html += f'<td style="padding: 8px; font-weight: normal; padding-left: 15px;">{prefixo}{descricao}{indicador_manual}</td>'
        html += f'<td style="padding: 8px; color: {cor_total}; text-align: center; font-weight: bold;" id="total_{input_id}">{int(total_exibido):,}</td>'
        
        # Adicionar valores mensais com IDs para JavaScript
        for mes in range(1, 13):
            valor_original = medias_recorrentes[descricao].get(mes, 0.0)
            
            # Usar valor manual se existir, senão usar o valor original
            if valor_manual_salvo:
                valor_mes_exibido = float(valor_manual_salvo)  # Usar valor manual diretamente
                estilo_manual = "background-color: rgba(40, 167, 69, 0.1); font-weight: bold;"
            else:
                valor_mes_exibido = valor_original
                estilo_manual = ""
            
            if valor_mes_exibido != 0:
                cor_valor = '#228B22' if valor_mes_exibido >= 0 else '#DC143C'
                html += f'<td style="padding: 8px; color: {cor_valor}; text-align: center; font-weight: bold; {estilo_manual}" id="mes_{mes}_{input_id}" data-original-value="{int(valor_original)}">{int(valor_mes_exibido):,}</td>'
            else:
                html += f'<td style="padding: 8px; text-align: center; color: #999; {estilo_manual}" id="mes_{mes}_{input_id}" data-original-value="{int(valor_original)}">-</td>'
        
        html += '</tr>'
        
        classificacao_anterior = classificacao_atual
    
    html += '</table></div>'
    
    
    # Adicionar JavaScript para atualizar valores manuais
    html += """
    <script>
    function updateMonthValues(inputId, originalTotal) {
        const inputElement = document.getElementById(inputId);
        const manualValue = parseFloat(inputElement.value) || 0;
        
        // Atualizar o total
        const totalElement = document.getElementById('total_' + inputId);
        if (manualValue > 0) {
            const color = manualValue >= 0 ? '#228B22' : '#DC143C';
            totalElement.innerHTML = manualValue.toLocaleString('pt-BR');
            totalElement.style.color = color;
        } else {
            const color = originalTotal >= 0 ? '#228B22' : '#DC143C';
            totalElement.innerHTML = originalTotal.toLocaleString('pt-BR');
            totalElement.style.color = color;
        }
        
        // Atualizar todos os 12 meses
        for (let mes = 1; mes <= 12; mes++) {
            const mesElement = document.getElementById('mes_' + mes + '_' + inputId);
            if (mesElement) {
                const originalValue = parseInt(mesElement.getAttribute('data-original-value'));
                
                if (manualValue > 0) {
                    // Usar valor manual diretamente (não dividir por 12)
                    const monthlyValue = Math.round(manualValue);
                    const color = monthlyValue >= 0 ? '#228B22' : '#DC143C';
                    mesElement.innerHTML = monthlyValue.toLocaleString('pt-BR');
                    mesElement.style.color = color;
                    mesElement.style.fontWeight = 'bold';
                } else {
                    // Voltar ao valor original
                    if (originalValue !== 0) {
                        const color = originalValue >= 0 ? '#228B22' : '#DC143C';
                        mesElement.innerHTML = originalValue.toLocaleString('pt-BR');
                        mesElement.style.color = color;
                        mesElement.style.fontWeight = 'bold';
                    } else {
                        mesElement.innerHTML = '-';
                        mesElement.style.color = '#999';
                        mesElement.style.fontWeight = 'normal';
                    }
                }
            }
        }
    }
    

    </script>
    """
    
    # Adicionar nota explicativa
    html += f"""
    <p style="margin-top: 15px; color: #666; font-style: italic; font-size: 14px;">
        📋 <strong>Nota:</strong> Esta projeção é baseada nas médias mensais dos itens marcados como recorrentes (RE).<br>
        💰 <strong>Saldo inicial:</strong> R$ {int(saldo_inicial_projecao):,} (último saldo calculado da tabela mensal atual)<br>
        🔧 <strong>Indicadores:</strong> Items com 🔧 têm valores manuais salvos no arquivo JSON
    </p>
    </div>
    """
    
    return html


def criar_tabela_mensal(dados_completos):
    """Cria tabela HTML organizada por descrição e meses"""
    
    # Ordem das classificações para ordenação
    ordem_classificacoes = [
        "RECEITAS",
        "EMPRÉSTIMOS",
        "REEMBOLSO", 
        "CONTA CORRENTE",
        "APLICAÇÃO FINANCEIRA",
        "TRANSFERENCIA ENTRE CONTAS",
        "DESPESAS",
        "IMPOSTOS",
        "FOLHA CLT",
        "FOLHA PJ",
        "ENCARGOS",
        "ADMINISTRATIVA",
        "ASSESSORIA JURIDICA",
        "ASSESSORIA CONTABIL",
        "DESPESAS FINANCEIRAS",
        "DESPESAS COMERCIAIS",
        "SOFTWARE",
        "PMT EMPRESTIMOS",
        "INVESTIMENTOS",
        "DESPESAS IMÓVEL",        
        "ADIANTAMENTO A FORNECEDORES",
        "NÃO CLASSIFICADO"
    ]
    
    # Definir subcategorias de DESPESAS
    subcategorias_despesas = [
        "IMPOSTOS",
        "FOLHA CLT",
        "FOLHA PJ",
        "ENCARGOS",
        "ADMINISTRATIVA",
        "ASSESSORIA JURIDICA",
        "ASSESSORIA CONTABIL",
        "DESPESAS FINANCEIRAS",
        "DESPESAS COMERCIAIS",
        "SOFTWARE",
        "PMT EMPRESTIMOS",
        "INVESTIMENTOS",
        "DESPESAS IMÓVEL",
        "ADIANTAMENTO A FORNECEDORES"
    ]
    
    # Carregar classificações para ordenação
    classificacoes = carregar_classificacoes()
    
    # Dicionário para armazenar os dados organizados
    tabela_dados = {}
    
    # Processar cada registro
    for registro in dados_completos:
        if len(registro) >= 3:
            data, descricao, valor = registro[0], registro[1], registro[2]
            
            # Extrair mês da data
            try:
                if isinstance(data, datetime):
                    mes = data.month
                elif isinstance(data, str):
                    # Tentar diferentes formatos de data
                    for formato in ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']:
                        try:
                            data_obj = datetime.strptime(data, formato)
                            mes = data_obj.month
                            break
                        except:
                            continue
                    else:
                        mes = 1  # Default para janeiro se não conseguir parsear
                else:
                    mes = 1  # Default
                
                # Converter valor para float
                if isinstance(valor, str):                
                    # Remove caracteres não numéricos exceto vírgula, ponto e sinal negativo
                    valor_limpo = re.sub(r'[^\d,.\-]', '', str(valor))
                    valor_limpo = valor_limpo.replace(',', '.')
                    
                    try:
                        valor_float = float(valor_limpo)
                        
                    except:
                        valor_float = 0.0
                        
                else:
                    valor_float = float(valor) if valor else 0.0                
                if descricao not in tabela_dados:                    
                    tabela_dados[descricao] = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
                
                # Somar valor ao mês correspondente                
                tabela_dados[descricao][mes] += valor_float                
                
            except Exception as e:
                print(f"Erro processando registro: {e}")
                continue
    
    # Criar HTML da tabela - versão simplificada
    meses_nomes = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                   'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    html = """
    <div style="max-height: 600px; overflow-y: auto;">
    <table border="1" style="width:100%; border-collapse: collapse;">
    <tr style="background-color: #f0f0f0;">
        <th style="text-align: left; padding: 8px; background-color: #f0f0f0;">Descrição</th>
        <th style="text-align: center; padding: 8px; background-color: #f0f0f0;">Total</th>
    """
    
    # Adicionar cabeçalhos dos meses
    for mes_nome in meses_nomes:
        html += f'<th style="text-align: center; padding: 8px;">{mes_nome}</th>'
    
    html += """
    </tr>
    """
    
    # Verificar se há dados para processar
    if not tabela_dados:
        return "<p>Nenhum dado encontrado para gerar a tabela mensal.</p>"
    
    # Definir classificações para linha de total especial
    classificacoes_receitas = [
        "RECEITAS",
        "EMPRÉSTIMOS",
        "REEMBOLSO", 
        "CONTA CORRENTE",
        "APLICAÇÃO FINANCEIRA",
        "TRANSFERENCIA ENTRE CONTAS"
    ]
    
    # Calcular totais das classificações de receitas
    def calcular_totais_receitas():
        totais_mes_receitas = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
        total_geral_receitas = 0.0
        
        for desc in tabela_dados.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            
            # Extrair apenas a classificação
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
                
            if classificacao_desc in classificacoes_receitas:
                dados_mes = tabela_dados[desc]
                for mes in range(1, 13):
                    totais_mes_receitas[mes] += dados_mes[mes]
                total_geral_receitas += sum(dados_mes.values())
        
        return totais_mes_receitas, total_geral_receitas
    
    # Calcular totais das despesas para o saldo bancário
    def calcular_totais_despesas():
        totais_mes_despesas = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
        total_geral_despesas = 0.0
        
        for desc in tabela_dados.keys():
            info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
            
            # Extrair apenas a classificação
            if isinstance(info_classificacao, dict):
                classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
            else:
                classificacao_desc = info_classificacao
                
            if classificacao_desc in subcategorias_despesas:
                dados_mes = tabela_dados[desc]
                for mes in range(1, 13):
                    totais_mes_despesas[mes] += dados_mes[mes]
                total_geral_despesas += sum(dados_mes.values())
        
        return totais_mes_despesas, total_geral_despesas
    
    # Calcular totais das receitas e despesas
    totais_mes_receitas, total_geral_receitas = calcular_totais_receitas()
    totais_mes_despesas, total_geral_despesas = calcular_totais_despesas()
    
    # Calcular saldo bancário cumulativo
    saldo_inicial = 272801.75
    saldo_acumulado = saldo_inicial
    saldos_mensais_cumulativos = {}
    
    # Calcular saldo acumulado mês a mês
    for mes in range(1, 13):
        movimento_mes = totais_mes_receitas[mes] + totais_mes_despesas[mes]  # receitas + despesas (despesas são negativas)
        saldo_acumulado += movimento_mes
        saldos_mensais_cumulativos[mes] = saldo_acumulado
    
    # Adicionar linha de saldo bancário (sem coluna total)
    html += f'<tr style="background-color: rgba(255, 215, 0, 0.4); font-weight: bold; border: 3px solid #FFD700;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 20px; font-weight: bold; color: #4169E1;">💰 SALDO BANCÁRIO</td>'
    html += f'<td style="padding: 8px; text-align: center; font-size: 16px; font-weight: bold; color: #666;">Saldo Inicial: {int(saldo_inicial):,}</td>'
    
    # Adicionar saldo acumulado de cada mês
    for mes in range(1, 13):
        saldo_mes_acumulado = saldos_mensais_cumulativos[mes]
        cor_saldo_mes = '#0066CC' if saldo_mes_acumulado >= 0 else '#DC143C'  # Azul para positivos, vermelho para negativos
        html += f'<td style="padding: 8px; color: {cor_saldo_mes}; text-align: center; font-weight: bold; font-size: 18px;">{int(saldo_mes_acumulado):,}</td>'
    
    html += '</tr>'
    
    # Adicionar linha de divisão estreita
    html += f'<tr style="height: 3px; border: none;"><td colspan="14" style="background-color: #666; height: 3px; border: none; padding: 0;"></td></tr>'
    
    # Adicionar linha RECEITA / DESPESAS (soma de receitas + despesas)
    total_receita_despesas = total_geral_receitas + total_geral_despesas
    html += f'<tr style="background-color: rgba(255, 165, 0, 0.4); font-weight: bold; border: 2px solid #FFA500;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold; color: #FF4500;">💰 RECEITA / DESPESAS</td>'
    
    # Calcular cor para o total (positivo = verde, negativo = vermelho)
    cor_total_receita_despesas = '#228B22' if total_receita_despesas >= 0 else '#DC143C'
    html += f'<td style="padding: 8px; color: {cor_total_receita_despesas}; text-align: center; font-size: 18px; font-weight: bold;">{int(total_receita_despesas):,}</td>'
    
    # Adicionar valores mensais (receitas + despesas por mês)
    for mes in range(1, 13):
        valor_mes_total = totais_mes_receitas[mes] + totais_mes_despesas[mes]
        if valor_mes_total != 0:
            cor_valor = '#228B22' if valor_mes_total >= 0 else '#DC143C'
            html += f'<td style="padding: 8px; color: {cor_valor}; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes_total):,}</td>'
        else:
            html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px; color: #FF4500;">-</td>'
    
    html += '</tr>'
    
    # Adicionar linha de divisão estreita
    html += f'<tr style="height: 3px; border: none;"><td colspan="14" style="background-color: #666; height: 3px; border: none; padding: 0;"></td></tr>'
    
    # Adicionar linha de total das receitas
    html += f'<tr style="background-color: rgba(144, 238, 144, 0.6); font-weight: bold; border: 2px solid #90EE90;">'
    html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold; color: #000080;">💰 RECEITAS/APLICAÇÕES/EMPRESTIMOS</td>'
    html += f'<td style="padding: 8px; color: #000080; text-align: center; font-size: 18px; font-weight: bold;">{int(total_geral_receitas):,}</td>'
    
    # Adicionar totais de cada mês para receitas
    for mes in range(1, 13):
        valor_mes = totais_mes_receitas[mes]
        if valor_mes != 0:
            html += f'<td style="padding: 8px; color: #000080; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes):,}</td>'
        else:
            html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px; color: #000080;">-</td>'
    
    html += '</tr>'
    
    # Adicionar linha de divisão mais grossa entre RECEITAS/APLICAÇÕES e dados detalhados
    html += f'<tr style="height: 8px; border: none;"><td colspan="14" style="background-color: #00CED1; height: 8px; border: none; padding: 0; border-top: 3px solid #008B8B; border-bottom: 2px solid #008B8B;"></td></tr>'
    
    # Função para obter índice da classificação para ordenação
    def obter_indice_classificacao(descricao):
        info_classificacao = classificacoes.get(str(descricao).strip(), "NÃO CLASSIFICADO")
        
        # Extrair apenas a classificação (formato novo ou antigo)
        if isinstance(info_classificacao, dict):
            classificacao = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
        else:
            classificacao = info_classificacao
            
        try:
            return ordem_classificacoes.index(classificacao)
        except ValueError:
            return len(ordem_classificacoes)  # Colocar no final se não encontrar
    
    # Ordenar descrições por classificação e depois alfabeticamente dentro da mesma classificação
    descricoes_ordenadas = sorted(tabela_dados.keys(), 
                                 key=lambda desc: (obter_indice_classificacao(desc), desc))
    
    # Calcular totais por classificação
    def calcular_totais_classificacao(classificacao):
        totais_mes = {i: 0.0 for i in range(1, 13)}  # Meses 1-12
        total_geral = 0.0
        
        # Se é DESPESAS, somar todas as subcategorias
        if classificacao == "DESPESAS":
            for desc in descricoes_ordenadas:
                info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
                
                # Extrair apenas a classificação
                if isinstance(info_classificacao, dict):
                    classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
                else:
                    classificacao_desc = info_classificacao
                    
                if classificacao_desc in subcategorias_despesas:
                    dados_mes = tabela_dados[desc]
                    for mes in range(1, 13):
                        totais_mes[mes] += dados_mes[mes]
                    total_geral += sum(dados_mes.values())
        else:
            # Lógica normal para outras classificações
            for desc in descricoes_ordenadas:
                info_classificacao = classificacoes.get(str(desc).strip(), "NÃO CLASSIFICADO")
                
                # Extrair apenas a classificação
                if isinstance(info_classificacao, dict):
                    classificacao_desc = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
                else:
                    classificacao_desc = info_classificacao
                    
                if classificacao_desc == classificacao:
                    dados_mes = tabela_dados[desc]
                    for mes in range(1, 13):
                        totais_mes[mes] += dados_mes[mes]
                    total_geral += sum(dados_mes.values())
        
        return totais_mes, total_geral

    # Adicionar dados de cada descrição ordenada com separadores por classificação
    classificacao_anterior = None
    despesas_ja_adicionada = False
    
    for descricao in descricoes_ordenadas:
        # Obter classificação atual
        info_classificacao = classificacoes.get(str(descricao).strip(), "NÃO CLASSIFICADO")
        
        # Extrair apenas a classificação
        if isinstance(info_classificacao, dict):
            classificacao_atual = info_classificacao.get('classificacao', 'NÃO CLASSIFICADO')
        else:
            classificacao_atual = info_classificacao
        
        # Se é uma subcategoria de DESPESAS e ainda não foi adicionada a linha DESPESAS
        if classificacao_atual in subcategorias_despesas and not despesas_ja_adicionada:
            # Adicionar linha DESPESAS primeiro
            totais_mes_despesas, total_despesas = calcular_totais_classificacao("DESPESAS")
            
            cor_total_despesas = 'red' if total_despesas < 0 else 'green'
            html += f'<tr style="background-color: rgba(64, 224, 208, 0.3); font-weight: bold; border: 2px solid rgba(64, 224, 208, 0.8);">'
            html += f'<td style="padding: 8px; text-align: center; font-size: 18px; font-weight: bold;">💰 DESPESAS MENSAIS</td>'
            html += f'<td style="padding: 8px; color: {cor_total_despesas}; text-align: center; font-size: 18px; font-weight: bold;">{int(total_despesas):,}</td>'
            
            # Adicionar totais de cada mês para DESPESAS
            for mes in range(1, 13):
                valor_mes = totais_mes_despesas[mes]
                if valor_mes != 0:
                    cor_valor = 'red' if valor_mes < 0 else 'green'
                    html += f'<td style="padding: 8px; color: {cor_valor}; text-align: center; font-weight: bold; font-size: 18px;">{int(valor_mes):,}</td>'
                else:
                    html += '<td style="padding: 8px; text-align: center; font-weight: bold; font-size: 18px;">-</td>'
            
            html += '</tr>'
            
            # Adicionar linha de divisão estreita após DESPESAS (TOTAL)
            html += f'<tr style="height: 3px; border: none;"><td colspan="14" style="background-color: #666; height: 3px; border: none; padding: 0;"></td></tr>'
            
            despesas_ja_adicionada = True
        
        # Se mudou a classificação, adicionar linha separadora com totais
        if classificacao_atual != classificacao_anterior:
            # Calcular totais para esta classificação
            totais_mes_classificacao, total_classificacao = calcular_totais_classificacao(classificacao_atual)
            
            # Linha da classificação com totais
            cor_total_classif = 'red' if total_classificacao < 0 else 'green'
            html += f'<tr style="background-color: #e8f4fd; font-weight: bold;">'
            html += f'<td style="padding: 8px; text-align: center;">📋 {classificacao_atual}</td>'
            html += f'<td style="padding: 8px; color: {cor_total_classif}; text-align: center;">{int(total_classificacao):,}</td>'
            
            # Adicionar totais de cada mês para a classificação
            for mes in range(1, 13):
                valor_mes = totais_mes_classificacao[mes]
                if valor_mes != 0:
                    cor_valor = 'red' if valor_mes < 0 else 'green'
                    html += f'<td style="padding: 8px; color: {cor_valor}; text-align: center; font-weight: bold;">{int(valor_mes):,}</td>'
                else:
                    html += '<td style="padding: 8px; text-align: center; font-weight: bold;">-</td>'
            
            html += '</tr>'
            classificacao_anterior = classificacao_atual
        
        dados_mes = tabela_dados[descricao]
        total_descricao = sum(dados_mes.values())
        
        cor_total = 'red' if total_descricao < 0 else 'green'
        html += f'<tr style="height: 30px;"><td style="padding: 4px 8px;">{descricao}</td>'
        html += f'<td style="padding: 4px 8px; color: {cor_total}; text-align: center;">{int(total_descricao):,}</td>'
        
        # Adicionar valores de cada mês
        for mes in range(1, 13):
            valor = dados_mes[mes]
            if valor != 0:
                cor_valor = 'red' if valor < 0 else 'black'
                html += f'<td style="padding: 4px 8px; color: {cor_valor}; text-align: center;">{int(valor):,}</td>'
            else:
                html += '<td style="padding: 4px 8px; text-align: center;">-</td>'
        
        html += '</tr>'
    
    html += """
    </table>
    """    
    return html


def main():
    # Configurar layout da página para usar toda a largura
    st.set_page_config(
        page_title="Processador de Extratos Excel - 4I Capital Ltda.",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Título centralizado, grande e em azul turquesa
    st.markdown("""
    <h1 style='text-align: center; color: #40E0D0; font-size: 3rem; font-weight: bold; margin-bottom: 1rem;'>
        Processador de Extratos Excel - Sicred e Bradesco 2025 - 4I Capital Ltda.
    </h1>
    """, unsafe_allow_html=True)
    
    # Verificar classificações sem recorrência na inicialização
    classificacoes_sem_recorrencia = verificar_classificacoes_sem_recorrencia()
    
    # Se existem classificações sem recorrência, mostrar alerta no topo
    if classificacoes_sem_recorrencia:
        st.error(f"🚨 **ATENÇÃO:** {len(classificacoes_sem_recorrencia)} classificação(ões) precisam ter a recorrência definida (RE/N_RE)!")
        st.warning("⬇️ **Role para baixo até a seção 'Classificação de Descrições' para definir as recorrências.**")
    
    arquivos = arquivos_disponiveis()    
    # Botão para processar arquivos
    dados_sicred = process_sicred_files(arquivos)
    arquivos = arquivos_disponiveis()
    dados_bradesco = process_bradesco_files(arquivos,dados_sicred)
    dados_completos = descricao(dados_bradesco)
    
    # Formulário de classificação das descrições
    if dados_completos:
        formulario_classificacao(dados_completos)    
    
    # Criar e exibir tabela HTML mensal
    if dados_completos:
        st.markdown("---")
        st.subheader("📊 Tabela Mensal por Descrição")
        st.markdown("")
        tabela_html = criar_tabela_mensal(dados_completos)
        st.markdown(tabela_html, unsafe_allow_html=True)
        
        # Criar e exibir tabela de fluxo futuro
        st.markdown("---")
        st.subheader("📈 Projeção de Fluxo Futuro - Próximos 12 Meses")
        st.markdown("")
        tabela_fluxo_futuro = criar_tabela_fluxo_futuro(dados_completos)
        st.markdown(tabela_fluxo_futuro, unsafe_allow_html=True)
        
        # Interface para gerenciar valores manuais
        st.markdown("---")
        st.subheader("💾 Gerenciar Valores Manuais de Projeção")
        
        # Mostrar status do arquivo
        arquivo_existe = os.path.exists('valores_manuais_projecao.json')
        if arquivo_existe:
            st.success("📄 Arquivo de valores manuais encontrado: `valores_manuais_projecao.json`")
        else:
            st.info("📄 Arquivo de valores manuais será criado ao salvar o primeiro valor")
        
        # Obter lista de descrições recorrentes para o selectbox
        medias_recorrentes = calcular_medias_recorrentes(dados_completos)
        descricoes_disponiveis = list(medias_recorrentes.keys()) if medias_recorrentes else []
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Carregar valores atuais para mostrar em um selectbox
            valores_atuais = carregar_valores_manuais()
            
            # Campo para adicionar/editar valor manual
            st.write("**Adicionar/Editar Valor Manual:**")
            
            if descricoes_disponiveis:
                # Selectbox com as descrições disponíveis
                descricao_selecionada = st.selectbox(
                    "Selecione a descrição:", 
                    [""] + descricoes_disponiveis, 
                    key="desc_select"
                )
                
                # Mostrar valor atual se existir
                if descricao_selecionada:
                    valor_atual = valores_atuais.get(descricao_selecionada, 0.0)
                    total_original = sum(medias_recorrentes[descricao_selecionada].values()) if descricao_selecionada in medias_recorrentes else 0
                    
                    if valor_atual != 0:
                        st.info(f"💾 Valor manual atual: R$ {valor_atual:,.2f} por mês (Total anual: R$ {valor_atual * 12:,.2f})")
                    else:
                        st.info(f"📊 Valor calculado atual: R$ {total_original/12:,.2f} por mês (Total anual: R$ {total_original:,.2f})")
                
                valor_input = st.number_input("Novo valor MENSAL:", value=0.0, key="valor_manual", help="Digite o valor que será usado TODOS OS MESES")
                
                col1a, col1b = st.columns(2)
                
                with col1a:
                    if st.button("💾 Salvar Valor", disabled=not descricao_selecionada):
                        if descricao_selecionada and valor_input != 0:
                            valores_atuais[descricao_selecionada] = valor_input
                            if salvar_valores_manuais(valores_atuais):
                                st.success(f"✅ Valor R$ {valor_input:,.2f} POR MÊS salvo para '{descricao_selecionada}'!")
                                st.success(f"📊 Total anual será: R$ {valor_input * 12:,.2f}")
                                st.rerun()
                            else:
                                st.error("❌ Erro ao salvar valor no arquivo JSON.")
                        else:
                            st.warning("⚠️ Selecione uma descrição e insira um valor diferente de zero.")
                
                with col1b:
                    if st.button("🗑️ Remover Valor", disabled=not descricao_selecionada or descricao_selecionada not in valores_atuais):
                        if descricao_selecionada in valores_atuais:
                            del valores_atuais[descricao_selecionada]
                            salvar_valores_manuais(valores_atuais)
                            st.success(f"✅ Valor removido para '{descricao_selecionada}'!")
                            st.rerun()
            else:
                st.warning("Nenhuma descrição recorrente encontrada.")
        
        with col2:
            # Mostrar valores salvos
            st.write("**Valores Manuais Salvos no Arquivo:**")
            if valores_atuais:
                for desc, valor in valores_atuais.items():
                    # Verificar se esta descrição está sendo usada na projeção
                    em_uso = desc in descricoes_disponiveis
                    status = "✅ Em uso" if em_uso else "⚠️ Não encontrada"
                    st.write(f"• **{desc}**: R$ {valor:,.2f}/mês | R$ {valor * 12:,.2f}/ano ({status})")
                
                st.write("---")
                
                # Mostrar exemplo do arquivo JSON
                with st.expander("📋 Ver conteúdo do arquivo JSON"):
                    st.json(valores_atuais)
                            
                if st.button("🗑️ Limpar Todos os Valores", type="secondary"):
                    if salvar_valores_manuais({}):
                        st.success("🗑️ Todos os valores foram removidos do arquivo!")
                        st.rerun()
            else:
                st.info("Nenhum valor manual salvo ainda no arquivo `valores_manuais_projecao.json`.")
                
            # Mostrar informações técnicas
            st.write("---")
            st.write("**ℹ️ Como funciona:**")
            st.write("1. Ao salvar um valor, ele é gravado no arquivo JSON")
            st.write("2. Na próxima atualização da tabela, o valor salvo substitui o calculado")
            st.write("3. O valor é usado diretamente em TODOS os 12 meses")
            st.write("4. Os totais são recalculados automaticamente")
        

  
def arquivos_disponiveis():
    # Informações sobre os arquivos na pasta    
    arquivos_dir = "\ArquivosExtratos"  
    
    if os.path.exists(arquivos_dir):
        arquivos = os.listdir(arquivos_dir)        
        arquivos_2025 = [arq for arq in arquivos if "2025" in arq]                 
        if arquivos_2025:            
            for arq in arquivos_2025:
                tipo = "Sicred" if "Sicred" in arq else "Bradesco" if "Bradesco" in arq else "Outro"
                
        else:
            st.write("Nenhum arquivo de 2025 encontrado.")
    else:
        st.error(f"Pasta {arquivos_dir} não encontrada!")
    path = arquivos_dir
    arq =[]
    for i in arquivos_2025:
        x = "ArquivosExtratos/"+i
        arq.append(x)    
    arquivos_dir = arq  
    return arquivos_dir

if __name__ == "__main__":

    main()
